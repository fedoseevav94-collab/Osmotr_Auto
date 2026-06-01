from __future__ import annotations

from dataclasses import dataclass
from difflib import get_close_matches
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.repository import InspectionRepository
from app.utils import normalize_plate


@dataclass(frozen=True)
class PlateHint:
    plate_normalized: str
    exact: bool


def _cell(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def read_vehicle_rows(path: Path) -> list[dict[str, str | None]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [_cell(value) for value in next(rows)]
    result: list[dict[str, str | None]] = []
    for values in rows:
        row = dict(zip(headers, values, strict=False))
        plate_raw = _cell(row.get("Номер"))
        if not plate_raw:
            continue
        plate_normalized = normalize_plate(plate_raw)
        if not plate_normalized:
            continue
        result.append(
            {
                "plate_raw": plate_raw,
                "plate_normalized": plate_normalized,
                "brand": _cell(row.get("Марка")),
                "model": _cell(row.get("Модель")),
                "status": _cell(row.get("Статус")),
                "source": str(path),
            }
        )
    return result


async def import_vehicle_registry(session: AsyncSession, path: Path) -> int:
    repo = InspectionRepository(session)
    count = 0
    for row in read_vehicle_rows(path):
        await repo.upsert_known_plate(**row)
        count += 1
    return count


async def plate_hint(session: AsyncSession, plate_normalized: str) -> PlateHint | None:
    repo = InspectionRepository(session)
    exact = await repo.get_known_plate(plate_normalized)
    if exact:
        return PlateHint(plate_normalized=plate_normalized, exact=True)

    known_values = await repo.list_known_plate_values()
    matches = get_close_matches(plate_normalized, known_values, n=1, cutoff=0.78)
    if matches:
        return PlateHint(plate_normalized=matches[0], exact=False)
    return None
