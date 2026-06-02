from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook

from app.constants import TIRE_TYPES
from app.models import DamageControlCase, InspectionSession

MOSCOW_TZ = ZoneInfo("Europe/Moscow")

SCORE_HEADERS = [
    "Гос номер",
    "Дата осмотра",
    "Тип осмотра",
    "Сотрудник осмотра",
    "Кузовные элементы",
    "Комментарий по кузову",
    "Техническое состояние",
    "Комментарий по техническому состоянию",
    "Оклейка",
    "Комментарий по оклейке",
    "Есть замечания водителя",
    "Комментарий по замечаниям водителя",
    "Тип резины",
    "Оценка резины",
    "Комментарий по резине",
    "Есть повреждения",
    "Описание повреждений",
    "Ссылка на сообщение в ФП",
]

HISTORY_HEADERS = SCORE_HEADERS[1:]

CHARGE_HEADERS = [
    "Гос номер",
    "Дата осмотра",
    "Тип осмотра",
    "Сотрудник осмотра",
    "Описание повреждений",
    "Сумма",
    "Тип оплаты",
    "Дата закрытия",
    "Ссылка на сообщение в ФП",
]


def period_bounds(period: str, now: datetime | None = None) -> tuple[datetime, datetime]:
    now = now or datetime.now(MOSCOW_TZ)
    today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    if period == "today":
        return today.replace(tzinfo=None), (today + timedelta(days=1)).replace(tzinfo=None)
    if period == "yesterday":
        start = today - timedelta(days=1)
        return start.replace(tzinfo=None), today.replace(tzinfo=None)
    if period == "week":
        start = today - timedelta(days=today.weekday())
        return start.replace(tzinfo=None), (start + timedelta(days=7)).replace(tzinfo=None)
    if period == "month":
        start = today.replace(day=1)
        if start.month == 12:
            end = start.replace(year=start.year + 1, month=1)
        else:
            end = start.replace(month=start.month + 1)
        return start.replace(tzinfo=None), end.replace(tzinfo=None)
    raise ValueError(f"Unknown period: {period}")


def fp_link(row: InspectionSession) -> str:
    if not row.fp_chat_id or not row.fp_message_id:
        return ""
    chat = str(row.fp_chat_id)
    if chat.startswith("-100"):
        chat = chat[4:]
    elif chat.startswith("100"):
        chat = chat[3:]
    return f"https://t.me/c/{chat}/{row.fp_message_id}"


def _base_values(row: InspectionSession) -> list[object]:
    return [
        row.completed_at.strftime("%d.%m.%Y %H:%M") if row.completed_at else "",
        row.scenario or "",
        f"@{row.telegram_username}" if row.telegram_username else row.telegram_name or "",
        row.body_score,
        row.body_comment or "",
        row.tech_score,
        row.tech_comment or "",
        row.wrap_score,
        row.wrap_comment or "",
        _yes_no(row.driver_has_remarks),
        row.driver_remarks_comment or "",
        TIRE_TYPES.get(row.tire_type or "", row.tire_type or ""),
        row.tire_score,
        row.tire_comment or "",
        "Да" if row.has_damage else "Нет",
        row.damage_description or "",
        fp_link(row),
    ]


def write_scores_xlsx(rows: list[InspectionSession], output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Оценки"
    ws.append(SCORE_HEADERS)
    for row in rows:
        ws.append([row.plate_normalized or row.plate_raw or "", *_base_values(row)])
    _autosize(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def write_history_xlsx(rows: list[InspectionSession], output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "История"
    ws.append(HISTORY_HEADERS)
    for row in rows:
        ws.append(_base_values(row))
    _autosize(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def write_problem_xlsx(rows: list[InspectionSession], output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Проблемные авто"
    ws.append([*SCORE_HEADERS, "Причина попадания"])
    for row in rows:
        ws.append([row.plate_normalized or row.plate_raw or "", *_base_values(row), problem_reason(row)])
    _autosize(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def write_charge_xlsx(rows: list[DamageControlCase], output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Списания"
    ws.append(CHARGE_HEADERS)
    for row in rows:
        inspection = row.inspection
        ws.append(
            [
                row.plate_normalized or inspection.plate_normalized or inspection.plate_raw or "",
                inspection.completed_at.strftime("%d.%m.%Y %H:%M") if inspection.completed_at else "",
                inspection.scenario or "",
                f"@{inspection.telegram_username}" if inspection.telegram_username else inspection.telegram_name or "",
                row.damage_description or inspection.damage_description or "",
                row.payment_amount or "",
                row.payment_type or "",
                row.closed_at.strftime("%d.%m.%Y %H:%M") if row.closed_at else "",
                fp_link(inspection),
            ]
        )
    _autosize(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def problem_reason(row: InspectionSession) -> str:
    reasons: list[str] = []
    if row.has_damage:
        reasons.append("повреждения")
    if row.body_score is not None and row.body_score < 4:
        reasons.append("кузов ниже 4")
    if row.tech_score is not None and row.tech_score < 4:
        reasons.append("техника ниже 4")
    if row.wrap_score is not None and row.wrap_score < 4:
        reasons.append("оклейка ниже 4")
    if row.tire_score is not None and row.tire_score < 4:
        reasons.append("резина ниже 4")
    return ", ".join(reasons)


def _yes_no(value: bool | None) -> str:
    if value is None:
        return ""
    return "Да" if value else "Нет"


def _autosize(ws) -> None:
    for column in ws.columns:
        letter = column[0].column_letter
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
        ws.column_dimensions[letter].width = width
